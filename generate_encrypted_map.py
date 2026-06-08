#!/usr/bin/env python3
"""
generate_encrypted_map.py

Reads student data from XLSX and a CSV of NISN→Google Drive file ID mappings,
then outputs two JSON files for the client-side secure download system:

  docs/encrypted_map.json   — AES-256-GCM encrypted payloads, keyed by SHA256(NISN)
  docs/student_index.json   — NISN_hash → Nama lookup (for "data ditemukan" feedback)

Usage:
  1. Prepare drive_files.csv with columns: NISN, drive_file_id
  2. Run:  python generate_encrypted_map.py
  3. Commit only the two output JSON files to GitHub.

Security:
  - Drive file IDs are NEVER stored in plain text
  - Each ID is encrypted with a key derived from the student's own NISN+":"+ddmmyyyy
  - The key uses PBKDF2 (100k iterations, SHA256)
  - Without knowing a specific student's credentials, the encrypted map is useless
"""

import hashlib
import json
import csv
import base64
import sys
from pathlib import Path

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
import openpyxl

# ---------------------------------------------------------------------------
# Configuration – do NOT hardcode plain Drive IDs here
# ---------------------------------------------------------------------------
SALT = b"mumtaza-tka-2026"
PBKDF2_ITERATIONS = 100_000
KEY_LENGTH = 32  # 32 bytes = AES-256

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "Data TKA SMP Mumtaza.xlsx"
CSV_PATH = BASE_DIR / "drive_files.csv"
OUTPUT_ENCRYPTED = BASE_DIR / "docs" / "encrypted_map.json"
OUTPUT_INDEX = BASE_DIR / "docs" / "student_index.json"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sha256_hex(s: str) -> str:
    """Return lowercase hex-encoded SHA256 of a string."""
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def derive_key(nisn: str, ddmmyyyy: str) -> bytes:
    """
    Derive a 32-byte AES-256 key from NISN + ":" + birth date.
    PBKDF2 with SHA-256, 100 000 iterations.
    """
    password = f"{nisn}:{ddmmyyyy}".encode("utf-8")
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=KEY_LENGTH,
        salt=SALT,
        iterations=PBKDF2_ITERATIONS,
    )
    return kdf.derive(password)


def encrypt_payload(key: bytes, payload: dict) -> dict:
    """
    Encrypt a dict payload with AES-256-GCM.
    Returns {"iv": b64, "ciphertext": b64, "tag": b64}.
    """
    aesgcm = AESGCM(key)
    # 96-bit IV / nonce
    iv = AESGCM.generate_nonce()
    plaintext = json.dumps(payload, separators=(",", ":")).encode("utf-8")
    # AESGCM.encrypt returns ciphertext + 16-byte tag appended
    ct_with_tag = aesgcm.encrypt(iv, plaintext, None)
    ct = ct_with_tag[:-16]      # ciphertext
    tag = ct_with_tag[-16:]     # authentication tag
    return {
        "iv": base64.b64encode(iv).decode(),
        "ciphertext": base64.b64encode(ct).decode(),
        "tag": base64.b64encode(tag).decode(),
    }


# ---------------------------------------------------------------------------
# Load XLSX data
# ---------------------------------------------------------------------------
def load_xlsx(path: Path) -> dict:
    """
    Return a dict keyed by NISN → {nama, ddmmyyyy, nomor_peserta, ...}
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    students = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        nisn = str(rec.get("NISN", "")).strip()
        if not nisn:
            continue
        # Normalise date
        dob = rec.get("ddmmyyyy")
        if isinstance(dob, (int, float)):
            dob = str(int(dob))
        else:
            dob = str(dob or "").strip()
        students[nisn] = {
            "nama": str(rec.get("Nama", "")).strip(),
            "ddmmyyyy": dob,
            "nomor_peserta": str(rec.get("Nomor_Peserta", "")).strip(),
        }
    return students


# ---------------------------------------------------------------------------
# Load drive_file_id CSV
# ---------------------------------------------------------------------------
def load_drive_csv(path: Path) -> dict:
    """Return dict NISN → drive_file_id from CSV."""
    mapping = {}
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            nisn = row.get("NISN", "").strip()
            fid = row.get("drive_file_id", "").strip()
            if nisn and fid:
                mapping[nisn] = fid
    return mapping


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    # ---- validate inputs ---------------------------------------------------
    if not XLSX_PATH.exists():
        print(f"[ERROR] XLSX file not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)
    if not CSV_PATH.exists():
        print(f"[ERROR] CSV file not found: {CSV_PATH}", file=sys.stderr)
        print("  Create drive_files.csv with columns: NISN,drive_file_id", file=sys.stderr)
        sys.exit(1)

    # ---- load data ---------------------------------------------------------
    print(f"Loading students from  {XLSX_PATH} …")
    students = load_xlsx(XLSX_PATH)
    print(f"  → {len(students)} students loaded")

    print(f"Loading Drive IDs from {CSV_PATH} …")
    drive_map = load_drive_csv(CSV_PATH)
    print(f"  → {len(drive_map)} Drive file IDs loaded")

    # ---- build encrypted map -----------------------------------------------
    encrypted_map = {}
    index_map = {}
    errors = []
    ok = 0

    for nisn, info in students.items():
        nama = info["nama"]
        dob = info["ddmmyyyy"]
        drive_id = drive_map.get(nisn)

        if not drive_id:
            errors.append(f"  ⚠  {nisn} ({nama}): no Drive file ID in CSV, skipped")
            continue
        if not dob:
            errors.append(f"  ⚠  {nisn} ({nama}): missing birth date, skipped")
            continue

        payload = {
            "fileId": drive_id,
            "nama": nama,
            "nomor_peserta": info["nomor_peserta"],
        }

        try:
            key = derive_key(nisn, dob)
            enc = encrypt_payload(key, payload)
        except Exception as e:
            errors.append(f"  ⚠  {nisn} ({nama}): encryption failed – {e}")
            continue

        nisn_hash = sha256_hex(nisn)
        encrypted_map[nisn_hash] = enc
        index_map[nisn_hash] = nama
        ok += 1

    # ---- write outputs -----------------------------------------------------
    OUTPUT_ENCRYPTED.parent.mkdir(parents=True, exist_ok=True)

    with open(OUTPUT_ENCRYPTED, "w", encoding="utf-8") as f:
        json.dump(encrypted_map, f, indent=2, ensure_ascii=False)
    print(f"\n✅ Encrypted map written → {OUTPUT_ENCRYPTED}")

    with open(OUTPUT_INDEX, "w", encoding="utf-8") as f:
        json.dump(index_map, f, indent=2, ensure_ascii=False)
    print(f"✅ Student index written  → {OUTPUT_INDEX}")

    # ---- summary -----------------------------------------------------------
    print(f"\n{'='*50}")
    print(f"  Total students in XLSX:  {len(students)}")
    print(f"  Drive IDs in CSV:        {len(drive_map)}")
    print(f"  Successfully encrypted:  {ok}")
    print(f"  Skipped / errors:        {len(errors)}")
    print(f"{'='*50}")
    if errors:
        print("\nErrors:")
        for e in errors:
            print(e)

    # Sanity-check: try decrypting the first entry
    if encrypted_map:
        print("\n🔍 Sanity-check: decrypting first entry …")
        try:
            first_hash = next(iter(encrypted_map))
            # Re-discover NISN from the hash (only possible because we know it)
            first_nisn = next(n for n, h in [(n, sha256_hex(n)) for n in students] if h == first_hash)
            first_dob = students[first_nisn]["ddmmyyyy"]
            k = derive_key(first_nisn, first_dob)
            enc = encrypted_map[first_hash]
            iv = base64.b64decode(enc["iv"])
            ct = base64.b64decode(enc["ciphertext"])
            tag = base64.b64decode(enc["tag"])
            aesgcm = AESGCM(k)
            decrypted = aesgcm.decrypt(iv, ct + tag, None)
            recovered = json.loads(decrypted.decode("utf-8"))
            print(f"  ✅ Decrypted OK: {recovered['nama']} → fileId={recovered['fileId'][:8]}…")
        except Exception as e:
            print(f"  ❌ Sanity-check FAILED: {e}", file=sys.stderr)
            sys.exit(1)


if __name__ == "__main__":
    main()
